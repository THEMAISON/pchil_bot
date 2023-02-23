from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
import schedule_data

short_names = {
    'None': '',
    'Архитектура вычислительных устройств ИС': 'Архитектура ВУ ИС',
    'Иностранный язык': 'Иностранный язык',
    'Компьютерная схемотехника': 'КC',
    'Объектно-ориентированное программирование': 'ООП',
    'Проектный модуль': 'Проектный модуль',
    'Русский язык делового общения': 'Русский язык',
    'Теория алгоритмов': 'Теория алгоритмов',
    'Теория вероятностей, вероятностные процессы и математическая статистика': 'Теория вероятностей',
    'Технологии проектной деятельности': 'ТПД',
    'Технологии создания программных продуктов': 'ТСПП',
    'Управление данными': 'Управление данными',
    'Элективные курсы по физической культуре и спорту': 'Элективы',
    'Понедельник': 'ПН',
    'Вторник': 'ВТ',
    'Среда': 'СР',
    'Четверг': 'ЧТ',
    'Пятница': 'ПТ',
    'Суббота': 'СБ',
    '№занятия': 'Пара',
    'подгруппа 1': 'Подгр. 1',
    'подгруппа 2': 'Подгр. 2'
}


def define_available_weeks():
    legacy_book = load_workbook(schedule_data.get_legacy_schedule_path(), read_only=True)

    for sheetname in legacy_book.sheetnames:
        # неделя 6(уч.н.24) -> неделя 6 -> 6
        week_name = str(sheetname).partition('(')[0].partition(' ')[2]
        if week_name.isdigit():
            schedule_data.available_weeks.append(int(week_name))


def improve_schedule_week():
    legacy_book = load_workbook(schedule_data.get_legacy_schedule_path(), read_only=True)
    legacy_sheet = legacy_book[f'неделя {schedule_data.week}(уч.н.{schedule_data.week + 18})']

    # создаем новый excel файл и применяем алгоритм улучшения расписания
    improve_book = Workbook()
    improve_sheet = improve_book.active

    # срез данных и копирование в новый файл
    improve_sheet = copy_data_week(legacy_sheet, improve_sheet)

    # форматирование и стилизация скопированных данных
    improve_sheet = style_data(improve_sheet)

    # сохраняем улучшенное расписание
    improve_book.save(schedule_data.get_improve_schedule_path())


def copy_data_week(legacy_sheet, improve_sheet):
    begin_column = 1
    begin_row = 4
    end_row = 54

    # определение начала среза по колонке
    for column in range(1, 200):
        cell = legacy_sheet.cell(row=4, column=column)
        group_name = str(cell.value).partition('Группа : ')[2].replace('\n', '').replace('/', '')

        if group_name == schedule_data.group_names.get(schedule_data.group):
            begin_column = column
            break

    # копирование данных по срезу
    for r in range(begin_row, end_row + 1):
        for c in range(begin_column, begin_column + 10):
            legacy_value = legacy_sheet.cell(row=r, column=c).value
            improve_sheet.cell(row=r - begin_row + 1, column=c - begin_column + 1, value=legacy_value)

    return improve_sheet


def style_data(improve_sheet):
    improve_sheet.title = f'{schedule_data.week} неделя'

    # устанавливаем ширину колонок
    improve_sheet.column_dimensions['A'].width = 5  # день недели
    improve_sheet.column_dimensions['B'].width = 10  # дата
    improve_sheet.column_dimensions['E'].width = 30  # занятие (1 подгруппа)
    improve_sheet.column_dimensions['H'].width = 30  # занятие (2 подгруппа)
    improve_sheet.column_dimensions['G'].width = 10  # аудитория (1 подгруппа)
    improve_sheet.column_dimensions['J'].width = 25  # аудитория (2 подгруппа)

    # выравнивание и сокращение всего текста
    for row in improve_sheet["A1:J51"]:
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell_value = str(cell.value)
            if not cell_value.isdigit():
                cell.value = cell_value.strip('\n')
                cell.value = do_short_names(cell_value)

    # выравнивание текста занятий и тип пары
    for row in range(4, 52):
        improve_sheet[f'F{row}'].value = str(improve_sheet[f'F{row}'].value).replace('\n', ' ').strip(' ')
        improve_sheet[f'G{row}'].value = str(improve_sheet[f'G{row}'].value).replace('\n', ' ').strip(' ')
        improve_sheet[f'I{row}'].value = str(improve_sheet[f'I{row}'].value).replace('\n', ' ').strip(' ')
        improve_sheet[f'J{row}'].value = str(improve_sheet[f'J{row}'].value).replace('\n', ' ').strip(' ')

    # исправление названия объединение общих пар
    for row in range(12, 52):
        solo_e = False
        solo_h = False
        pair_name_e = improve_sheet[f'E{row}'].value
        pair_name_h = improve_sheet[f'H{row}'].value

        if pair_name_e or pair_name_h:
            improve_sheet[f'C{row}'].fill = set_fill_color('DAEEF3')
            improve_sheet[f'D{row}'].fill = set_fill_color('DAEEF3')

        if pair_name_e:
            pair_name_e_text, solo_e = is_solo_pair(str(pair_name_e))
            improve_sheet[f'E{row}'].value = do_short_names(pair_name_e_text)

            # выделяем существующую пару
            improve_sheet[f'E{row}'].fill = set_fill_color('DAEEF3')
            improve_sheet[f'F{row}'].fill = set_fill_color('DAEEF3')
            improve_sheet[f'G{row}'].fill = set_fill_color('DAEEF3')

        if pair_name_h:
            pair_name_h_text, solo_h = is_solo_pair(str(pair_name_h))
            improve_sheet[f'H{row}'].value = do_short_names(pair_name_h_text)

            # выделяем существующую пару
            improve_sheet[f'H{row}'].fill = set_fill_color('DAEEF3')
            improve_sheet[f'I{row}'].fill = set_fill_color('DAEEF3')
            improve_sheet[f'J{row}'].fill = set_fill_color('DAEEF3')

        if (not solo_e) and (not solo_h) and pair_name_e:
            improve_sheet.merge_cells(f'E{row}:H{row}')

            # исправление типа пары англ
            if improve_sheet[f'E{row}'].value == do_short_names('Иностранный язык'):
                improve_sheet[f'I{row}'].value = str(improve_sheet[f'I{row}'].value).partition(' ')[0]

            improve_sheet[f'H{row}'].fill = set_fill_color('DAEEF3')
            improve_sheet[f'I{row}'].fill = set_fill_color('DAEEF3')
            improve_sheet[f'J{row}'].fill = set_fill_color('DAEEF3')

    # форматирование блока ПУЛа
    for row in range(4, 11):
        if improve_sheet[f'E{row}'].value:
            improve_sheet[f'E{row}'].value = fix_pul_pair_name(str(improve_sheet[f'E{row}'].value))
            improve_sheet.merge_cells(f'E{row}:H{row}')

            improve_sheet[f'E{row}'].fill = set_fill_color('DAEEF3')
            improve_sheet[f'C{row}'].fill = set_fill_color('DAEEF3')
            improve_sheet[f'D{row}'].fill = set_fill_color('DAEEF3')
            improve_sheet[f'I{row}'].fill = set_fill_color('DAEEF3')
            improve_sheet[f'J{row}'].fill = set_fill_color('DAEEF3')

    # форматирование шапки
    improve_sheet['A1'].fill = set_fill_color('215967')

    # устанавливаем жирный текст на шапке
    for row in improve_sheet["A1:J3"]:
        for cell in row:
            cell.font = Font(bold=True)

    improve_sheet.merge_cells('A1:J1')  # название группы
    improve_sheet.merge_cells('A2:A3')  # день
    improve_sheet.merge_cells('B2:B3')  # дата
    improve_sheet.merge_cells('C2:C3')  # №занятия
    improve_sheet.merge_cells('D2:D3')  # время
    improve_sheet.merge_cells('E2:G2')  # подгруппа 1
    improve_sheet.merge_cells('H2:J2')  # подгруппа 1

    for row in range(4, 52, 8):
        improve_sheet.merge_cells(f'A{row}:A{row + 7}')  # данные дня
        improve_sheet.merge_cells(f'B{row}:B{row + 7}')  # данные даты

    return improve_sheet


def set_fill_color(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


def fix_pul_pair_name(pair_name: str) -> str:
    legacy_names_group = pair_name.split('\n')
    improved_group_names = []

    for name in legacy_names_group:
        if 'Уч.гр. ' in name:
            name = name[name.find('Уч.гр. '):].partition(' ')[2].partition(')')[0]
        elif ')' in name:
            name = name.partition('(')[0].strip(' ')
        else:
            name = name.partition(',')[0]
        improved_group_names.append(name)

    return ' '.join(improved_group_names)


def is_solo_pair(pair_name: str):
    if 'подгр.:' in pair_name:
        return pair_name.partition(',')[2].partition(',')[0], True

    return pair_name.partition(',')[0], False


def do_short_names(name):
    if name in short_names.keys():
        return short_names.get(name)
    return name


def define_days_week_messages():
    improve_book = load_workbook(schedule_data.get_improve_schedule_path(), read_only=True)
    improve_sheet = improve_book.active

    day_count = 1

    for row in range(4, 52, 8):
        day_week_message = f'📖 *Расписание. {schedule_data.group_names.get(schedule_data.group)}. ' \
                           f'{schedule_data.days_week_names[day_count]}. ' \
                           f'' + str(improve_sheet[f'B{row}'].value) + '*'

        for pair_num in range(row, row + 8):
            time_block = improve_sheet[f'D{pair_num}'].value
            s1group = improve_sheet[f'E{pair_num}'].value
            s1type_pair = improve_sheet[f'F{pair_num}'].value
            s1pair_room = improve_sheet[f'G{pair_num}'].value
            s2group = improve_sheet[f'H{pair_num}'].value
            s2type_pair = improve_sheet[f'I{pair_num}'].value
            s2pair_room = improve_sheet[f'J{pair_num}'].value

            if (not s1group) and (not s2group):
                continue

            pair_text = f'🔹*{pair_num - row + 1} Пара - {time_block}*'

            if s1group and (not s2group) and (not s1type_pair):
                pair_text += f'\n{s1group}, *{s2type_pair}*'
                pair_text += f', *{s2pair_room}*' if s2pair_room else ''
            else:
                if s1group:
                    pair_text += f'\n{s1group}'
                    pair_text += f', *{s1type_pair}*' if s1type_pair else ''
                    pair_text += f', *{s1pair_room}*' if s1pair_room else ''
                    pair_text += ' *(1 подгр.)*'
                if s2group:
                    pair_text += f'\n{s2group}'
                    pair_text += f', *{s2type_pair}*' if s2type_pair else ''
                    pair_text += f', *{s2pair_room}*' if s2pair_room else ''
                    pair_text += ' *(2 подгр.)*'

            day_week_message += '\n' + pair_text

        schedule_data.days_week_messages[day_count] = day_week_message
        day_count += 1
